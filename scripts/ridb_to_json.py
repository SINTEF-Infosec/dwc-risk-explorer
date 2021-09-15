from openpyxl import load_workbook
import json

RIDB_FILE = 'data/ridb.xlsx'
OUTPUT_FILE = "public/resources/events.json"

print("[+] Loading RIDB from %s" % RIDB_FILE)
wb = load_workbook(filename = RIDB_FILE)
wb.active = 3 # Selection of the database sheet
ws = wb.active


def read_event(i):
    print("[+] Handling event %d" % i)
    evt_id = str(ws["A{}".format(2 + i)].value)
    evt_type_of_source = ws["B{}".format(2 + i)].value
    evt_type_of_threat = ws["C{}".format(2 + i)].value
    evt_type_of_event = ws["D{}".format(2 + i)].value
    evt_supporting_asset = ws["E{}".format(2 + i)].value
    evt_composite_asset = ws["F{}".format(2 + i)].value
    evt_primary_asset = ws["G{}".format(2 + i)].value
    evt_consequence = ws["H{}".format(2 + i)].value
    evt_description = ws["I{}".format(2 + i)].value
    evt_example = ws["L{}".format(2 + i)].value
    evt_measures = [m.strip() for m in str(ws["M{}".format(2 + i)].value).split(",")]

    event = {}
    event["id"] = evt_id
    event["type_of_source"] = evt_type_of_source
    event["type_of_threat"] = evt_type_of_threat
    event["type_of_event"] = evt_type_of_event
    event["supporting_asset"] = evt_supporting_asset
    event["composite_asset"] = evt_composite_asset
    event["primary_asset"] = evt_primary_asset
    event["consequence"] = evt_consequence
    event["description"] = "%s generates a %s threat causing a %s of the %s of the %s which affects %s and might lead to a %s issue" % (evt_type_of_source, evt_type_of_threat, evt_type_of_event, evt_supporting_asset, evt_composite_asset, evt_primary_asset, evt_consequence)
    event["example"] = evt_example
    event["measures"] = evt_measures

    return event

events = []
keep_reading = True
k = 0
while keep_reading:
    m = read_event(k)
    if m["type_of_source"] == "" or m["type_of_source"] == None:
        break
    events.append(m)
    k += 1

with open(OUTPUT_FILE, "w") as of:
    json.dump(events, of)
