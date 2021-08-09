from openpyxl import load_workbook
import json

RRMD_SPAN = 24

wb = load_workbook(filename = 'data/ridb.xlsx')
wb.active = 3 # Selection of the database sheet
ws = wb.active

wb2 = load_workbook(filename = 'data/rrmd.xlsx')
wb2.active = 3 
ws2 = wb2.active


def init_measures(m_dict, s):
    for k in range(s):
        m_dict["{}".format(k)] = {
            "type_of_source": set(),
            "type_of_threat": set(),
            "type_of_event": set(),
            "specific_asset": set(),
            "type_of_asset": set(),
            "consequence": set(),
        }

def read_event_and_update_measures(i, measures):
    print("[+] Handling event %d" % i)
    evt_id = str(ws["A{}".format(2 + i)].value)
    evt_type_of_source = ws["B{}".format(2 + i)].value
    evt_type_of_threat = ws["C{}".format(2 + i)].value
    evt_type_of_event = ws["D{}".format(2 + i)].value
    evt_supporting_asset = ws["E{}".format(2 + i)].value
    evt_composite_asset = ws["F{}".format(2 + i)].value
    evt_primary_asset = ws["G{}".format(2 + i)].value
    evt_consequence = ws["H{}".format(2 + i)].value
    evt_description = ws["I{}".format(2 + i)].value
    evt_measures = [m.strip() for m in str(ws["M{}".format(2 + i)].value).split(",")]

    for measure_id in evt_measures:
        measures[measure_id]["type_of_source"].add(normalize_text(evt_type_of_source))
        measures[measure_id]["type_of_threat"].add(normalize_text(evt_type_of_threat))
        measures[measure_id]["type_of_event"].add(normalize_text(evt_type_of_event))
        measures[measure_id]["specific_asset"].add(normalize_text(evt_supporting_asset))
        measures[measure_id]["specific_asset"].add(normalize_text(evt_composite_asset))
        measures[measure_id]["type_of_asset"].add(normalize_text(evt_primary_asset))
        measures[measure_id]["consequence"].add(normalize_text(evt_consequence))


def update_rrmd(measures):
    s_c = 0
    u_c = 0

    for m_id, values in measures.items():
        if values["type_of_source"] == set():
            print("[-] Skipping measure %s" % m_id)
            s_c += 1
            continue
        
        start_row = 2 + RRMD_SPAN * (int(m_id) - 1)
        cell = "B{}".format(start_row)

        m_name = ws2[cell].value
        print("[+] Updating %s" % m_name)
       
        # Cleaning first
        clean_cell("E", start_row)
        clean_cell("F", start_row)
        clean_cell("G", start_row)
        clean_cell("H", start_row)
        clean_cell("I", start_row)
        clean_cell("J", start_row)

        add_list_cell("E", start_row, values["type_of_source"])
        add_list_cell("F", start_row, values["type_of_threat"])
        add_list_cell("G", start_row, values["type_of_event"])
        add_list_cell("H", start_row, values["specific_asset"])
        add_list_cell("I", start_row, values["type_of_asset"])
        add_list_cell("J", start_row, values["consequence"])

        u_c += 1

    print("Done with the update: %s updated, %s skipped" % (u_c, s_c))

def clean_cell(column, start_row):
    for k in range(RRMD_SPAN):
        c = "{}{}".format(column, start_row + k)
        ws2[c] = ""

def add_list_cell(column, start_row, l):
    for idx, elt in enumerate(l):
        if idx >= RRMD_SPAN:
            print("********** TOO MANY ITEMS IN THE LIST, SKIPPING **********")
            break
        ws2["{}{}".format(column, start_row + idx)] = elt

def normalize_text(t):
    return t.lower().capitalize()

measures = {}

init_measures(measures, 75)

for k in range(73):
    read_event_and_update_measures(k, measures)


c = 0
d = 0
for mid, values in measures.items():
    if len(values["type_of_source"]) == 0:
        c += 1
    else:
        d += 1

print("%d measures not used, %d used" % (c, d))

update_rrmd(measures)

wb2.save(filename='data/rrmd.xlsx')


# rrm = [getMeasure(k) for k in range(65)]

# with open('rrmdb.json', 'w') as outfile:
#     json.dump(rrm, outfile)
